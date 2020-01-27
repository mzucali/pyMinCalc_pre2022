'''
Created on Mar 28, 2019

@author: miki
'''
#import START

#from openpyxl.writer.excel import ExcelWriter


import os
from collections import OrderedDict

import mineral_constants
# from openpyxl import worksheet, Workbook
import pandas as pd
# import xlwt
import xlsxwriter
from openpyxl import load_workbook
from xlrd import open_workbook

# import ntpath
# import csv
# import pprint
# from copy import deepcopy
# import openpyxl
# import xlrd
# from formula import headers, data_input_Ox_dict_list
# from xlutils.copy import copy
# import END


row_oxides = 0
col_oxides = 0
#workbook = xlsxwriter.Workbook()


    
def readFILE_E_ESTRAI_DATI_MA_CONTROLLA_MINLABEL_SET_OX(file_Input_XLSX):

    data_input_Ox_dict_list = []
    fin = file_Input_XLSX
    wb = open_workbook(fin)
    #wb = open_workbook(file_Input_XLSX)

    for s in wb.sheets():
        print ('\n\nSheet: {0}, columns = {1}, rows = {2} '.format(s.name,s.ncols, s.nrows))
        for riga in range(1):
            #global headers
            headers = []
            for colo in range(s.ncols):
                headers.append(str(s.cell(riga,colo).value))
            #print ('headers read in Sheet: ', s.name)
            #print("SAMPLE: ",headers[0])
            #headers[0] = 'sample'
            headers[1] ='mineral'
            #print("SAMPLE: ",headers[0])
            #print("MINERAL ", headers[1])
            print ('\t'.join(i for i in headers))

        print ()

    print ("LIST ALL ANALYSES and STORE in dict_list")
    for s in wb.sheets():
        print ('\nSheet:',s.name)
        print()
        for rows in range(1, s.nrows):
            values = []
            for cols in range(s.ncols):
                values.append(s.cell(rows,cols).value)
            print ("Analysis = %s" %rows)
            print ("Headers:\t" + '\t'.join([str(i) for i in headers]))
            print ("Values:\t" +  '\t'.join([str(i) for i in values]))
            #print ("\n")
            
            
                        
            data_input_Ox_dict = OrderedDict(zip(headers, values))
            print("INOUTFILE data_input_Ox_dict = ",data_input_Ox_dict) 
#             for k,v in data_input_Ox_dict.items():
#                 print("k e v: ", k, v)
            data_input_Ox_dict2 = OrderedDict(zip(headers, values))
            data_input_Ox_dict2 = changeMineralLabels(data_input_Ox_dict)
            print("INOUTFILE data_input_Ox_dict CHANGED MIN LABEL= ",data_input_Ox_dict2) 
            
            #data_input_Ox_dict_list = []
            #print("from READEXCEL2 added analysis = ",data_input_Ox_dict)
            #print()
            data_input_Ox_dict_list.append(data_input_Ox_dict2)

    print()

    wb.release_resources()
    
    ## SET mineral to 'mineral' and sample to 'Sample'
    ## CHECK MINERAL LABEL and change to unique LABEL=>>DONE
    ## 
    
    print("INOUTdata_input_Ox_dict_list..",data_input_Ox_dict_list)
    
    return data_input_Ox_dict_list



def changeMineralLabels(cats_dict):

    if cats_dict['mineral'] in mineral_constants.mineral_labels.keys():
        v = cats_dict['mineral']
        print("changeALLKeys found: ", v)

        print(mineral_constants.mineral_labels[v])
        new_value = mineral_constants.mineral_labels[v]
        cats_dict['mineral']=new_value

    return cats_dict

         
   
    
def removeEXT(file_name_path):
    filename = os.path.splitext(file_name_path)[0]
    return filename    




def write_out_base_data(recalc_data_oxides_cats_OX__list, inputfile_path):
 
    global wbook
    global wsheet

 
    print ("FILE INPUT ..."+ inputfile_path)
    file_output = removeEXT(inputfile_path) + '_OUT.xlsx'
    print ("FILE OUTPUT ..."+ file_output)
    print ("WRITING HEADERS to FileOUT " + str(file_output))

    wbook = xlsxwriter.Workbook(file_output)
    wsheet = wbook.add_worksheet("data")

 
    row = 0
    col = 0
    print("DATASET write headers from oxides_dict_list...",recalc_data_oxides_cats_OX__list)
    print("DATASET write headers from self.data_input_OX_list",recalc_data_oxides_cats_OX__list)
    for k in recalc_data_oxides_cats_OX__list[0]:
        wsheet.write(row, col, k)
        row += 1
    row = 0
    col += 1

    print ("column = " + str(col))
    print ("row = " + str(row))


#     global headers_written
#     headers_written = True
    
    global row_oxides
    row_oxides = row
    global col_oxides
    col_oxides = col
    
    
    
    for mineral in recalc_data_oxides_cats_OX__list:
        for k,v in mineral.items():
            wsheet.write(row, col, v)
            row += 1
        col += 1
        row = 0
        
    print("GLOBAL WNOOK: ", wbook) 
    wbook.close()
    transpose_excel(file_output)
    
    
    
    return file_output


def transpose_excel(fileOUT):
    df = pd.read_excel(fileOUT, sheet_name="data")
    #df.as_matrix()
    df.values
    df.transpose()
    print ("table")
    print (str(df))
    print ("transpose")
    print (str(df.transpose()))
    writer1 = pd.ExcelWriter(fileOUT)
    df.to_excel(writer1, sheet_name='TAB', index=False)
    df.transpose().to_excel(writer1,sheet_name='APPEND', header = False)
    writer1.save()
 
    return
 
 
def write_out_data_by_mineral_with_specific_sites2(recalc_data_oxides_cats_OX_by_mineral_list, fileOUT):
    
    

    global wb2
    wb2 = load_workbook(fileOUT)
    #print("WB2 = ", wb2)
    
    
    for k,v in recalc_data_oxides_cats_OX_by_mineral_list.items():
        #print("k and v: ", k, v)
        #print("WB2 = ", wb2)

        global sheet
        sheet = wb2.create_sheet(k)

        global row
        global col
        row = 1
        col = 1   
        
        for l in v[0]:
            
            sheet.cell(row=row, column=col).value = l
            row += 1


        row = 1
        col += 1
        
        for mineral in v:
            for k,v in mineral.items():
                sheet.cell(row=row, column=col).value = v
                row += 1
            col += 1
            row = 1    
    
    
    wb2.save(fileOUT)
    
    

def write_out_data_by_mineral_with_specific_sites(recalc_data_oxides_cats_OX_by_mineral_list, fileOUT):
    print("INOUTFILE=>write_out_data_by_mineral_with_specific_sites(recalc_data_oxides_cats_OX_by_mineral_list)")
    
    for k,v in recalc_data_oxides_cats_OX_by_mineral_list.items():
        print("k and v: ", k, v)
        
        print ("FILE INPUT ..."+ fileOUT)
        file_output = removeEXT(fileOUT) +'_'+ k+'.xlsx'
        print ("FILE OUTPUT ..."+ file_output)

        wbook = xlsxwriter.Workbook(file_output)
        wsheet = wbook.add_worksheet(k)
        
        global row
        global col
        row = 0
        col = 0

        for l in v[0]:
            
            wsheet.write(row, col, l)
            row += 1
        row = 0
        col += 1
 
 
        for mineral in v:
            for k,v in mineral.items():
                wsheet.write(row, col, v)
                row += 1
            col += 1
            row = 0
     
        
     
    
    
        wbook.close()
    
    return


def writeAX_formatted_input(recalc_data_oxides_cats_OX_by_mineral_list, fileOUT):
    ##to be implemented
    '''
    write headers
    read by mineral
    write by mineral
        sample mineral (check AX)
        new line
        write oxides
    
    '''
    return


    
    
    
    
    